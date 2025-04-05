from openai import OpenAI
from openpyxl import Workbook
import requests
import pdfplumber
from easyofd import OFD
import os
import json
import re

class RecToForm:
    def __init__(self, in_path, out_path, message, api_key, base_url):
        self.in_path = in_path
        self.out_path = out_path
        self.message = message
        self.api_key = api_key
        self.base_url = base_url


    def getfile(self):
        '''
        获取指定文件列表中各个文件, 首先把ofd格式转换为pdf格式, 再对pdf进行信息提取出字符串
        :return:
                self.pdf_informations:[str]
        '''
        self.pdf_informations:[str] = []  #pdf_informations列表中每个字符串为每张发票的具体信息
        self.newfiles_path:[str] = [] #暂时存储新生成的pdf文件

        for files_path in os.listdir(self.in_path):
            file_path = os.path.join(self.in_path, files_path)
            file_name, file_extension = os.path.splitext(files_path)
            if file_extension == ".ofd":
                ofd = OFD()
                ofd.read(file_path, fmt="path")
                pdf_bytes = ofd.to_pdf()
                newfile_path = self.in_path + '/' + file_name + ".pdf"
                self.newfiles_path.append(newfile_path)
                with open(newfile_path, "wb") as f:
                    f.write(pdf_bytes)
                ofd.disposal()

        for files_path in os.listdir(self.in_path):
            file_path = os.path.join(self.in_path, files_path)
            file_name, file_extension = os.path.splitext(files_path)
            if file_extension == ".pdf":
                with pdfplumber.open(file_path) as pdf:
                    for i, page in enumerate(pdf.pages):
                        text = page.extract_text()
                        self.pdf_informations.append(text)

        #删除暂时新生成的pdf文件
        for newfile_path in self.newfiles_path:
            os.remove(newfile_path)

    def recognize(self):
        """
        将self.pdf_informations:[str]列表信息提取, 传递给deepseek分析, 本地正则数据处理为可用字典
        :return:
                self.dataframes:[dict]
        """
        self.dataframes:[dict] = [] #列表中的每个元素为发票信息处理后的字典格式
        idx = 1
        client = OpenAI(api_key=self.api_key, base_url=self.base_url)
        for pdf_information in self.pdf_informations:
            response = client.chat.completions.create(
                model = "deepseek-chat",
                messages = [
                    {"role": "assistant", "content": self.message},
                    {"role": "assistant", "content": pdf_information}
                ]
            )
            information_modified = re.sub(r"```(?:python)?\n?", "", response.choices[0].message.content).strip()
            dict_modified = json.loads(information_modified)
            self.dataframes.append(dict_modified)
            print(f"第{idx}张发票已成功上传并分析")
            idx += 1

    def fill(self):
        """
        使用self.dataframes填写excel表格
        """
        wb = Workbook()
        ws = wb.active
        heads = ["序号", "发票代码", "发票号", "发票金额"]
        for i in range(0, len(heads)):
            ws.cell(1, i+1, heads[i])
        idx_row, idx_col = 2, 2
        for dataframe in self.dataframes: #self.dataframes为请求api后返回的列表
            ws.cell(idx_row, 1, idx_row-1)
            for data in dataframe.values():
                if data is None:
                    data = "None"
                ws.cell(idx_row, idx_col, data)
                idx_col += 1
            idx_col = 2
            idx_row += 1
        wb.save(self.out_path)

        print("表格已填写完成, 请查看")

def hint():
    print("作者：Yunxi_Zhu")
    print("----------------发票信息自助填表---------------")
    print("在使用前, 请预先创建一个文件夹, 并在里边创建“发票”的文件夹(内含格式为pdf, ofd的发票文件)\n同时, 将本程序至于其同级目录下")
    flag = input("请确保你已完成上述操作[y/n]：")
    return flag == "y"


if __name__ == "__main__":
    in_path = r'发票'
    out_path = r'发票信息.xlsx'
    api_key = "your_api_key"
    base_url = "https://api.deepseek.com/v1"
    message = "分析下列的发票文件，提取发票代码、发票号码、发票金额这些信息。只需要发票代码、发票号码、发票金额（价税合计中的小写金额），注意：发票代码（一定是该关键字，不要误判）与发票号码不一样，若没有数据，则将发票代码字段填入与发票号码一样的值(也是字符串)，最后以python的字典的形式返回（只需要字典，其余多余字符串不需要）\n"

    flag = hint()
    if flag:
        rtf = RecToForm(in_path, out_path, message, api_key, base_url)

        rtf.getfile()
        rtf.recognize()
        rtf.fill()